"""
CNKI 论文引文获取工具 v3 — 拟人化版本

核心策略（规避 CNKI 机器人校验）：
1. 使用系统真实 Chrome（channel="chrome"），指纹与真人浏览器一致
2. 持久化 Profile（.chrome_profile），保留 Cookie / 登录态 / 浏览器指纹
3. 移除 Playwright 自动化标志（--enable-automation）
4. 拟人行为模拟：贝塞尔曲线鼠标轨迹、逐字随机延迟输入、随机滚动、随机停顿
5. playwright-stealth 隐藏 navigator.webdriver 等自动化特征

用法：
  python cnki_citation.py "论文标题"                 # 单篇
  python cnki_citation.py -f titles.txt             # 批量（每行一个标题）
  python cnki_citation.py --connect "论文标题"       # 连接已打开的 Chrome
"""

import asyncio
import re
import json
import sys
import random
import argparse
import urllib.request
import urllib.parse
import ssl
import certifi
from pathlib import Path
from typing import Optional, List, Tuple
from datetime import datetime

from playwright.async_api import async_playwright, Page
from playwright_stealth import Stealth
import openpyxl
from openpyxl.utils import column_index_from_string


# 基准目录：
# - 脚本态：与脚本同目录（开发调试，Profile/输出落在项目里）
# - 冻结态（exe）：_MEI 临时目录每次启动都不同，放那里会导致持久 Profile/输出丢失。
#   改用稳定的用户目录 C:\Users\<user>\.cnki_citation，保证 Cookie/登录态/指纹跨运行保留。
if getattr(sys, "frozen", False):
    BASE_DIR = Path.home() / ".cnki_citation"
else:
    BASE_DIR = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / "output"
PROFILE_DIR = BASE_DIR / ".chrome_profile"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
PROFILE_DIR.mkdir(parents=True, exist_ok=True)

CNKI_HOME = "https://www.cnki.net"


# ── 版本与更新检测（Gitee Releases）──
# 发布流程：在 Gitee 仓库「发行版」页创建 Release（tag 如 v1.0.1），工具启动时比对最新 tag
APP_VERSION = "1.0.4"
GITEE_OWNER = "sulele"           # Gitee 用户名（与 build_release.py / .github/workflows/sync_gitee.py 保持一致）
GITEE_REPO = "cnki-citation"      # Gitee 仓库名
# 私有仓库必须带 token 才能访问 Releases API；公开仓库留空即可。
# 注意：此 token 会被打包进 exe，任何人可反编译提取，故务必用「只读」令牌，
# 且只授予本仓库，切勿用有写权限的令牌（发布用单独的写令牌，见 build_release.py）。
# 用于检测更新的 Gitee 令牌（当前取自 .release_token 的写令牌；
# ⚠️ 安全提示：写令牌嵌进 exe 后，任何能反编译 exe 的人都能拿到它并拥有该仓库的写权限。
# 生产环境强烈建议改用「只读」PAT（Gitee→设置→私人令牌，仅勾 projects/releases 只读）。
GITEE_TOKEN = "7fdea635702f2c7d1005de1620476aa1"
GITEE_API = f"https://gitee.com/api/v5/repos/{GITEE_OWNER}/{GITEE_REPO}/releases/latest"


def _norm_ver(v: str) -> Tuple[int, int, int]:
    """把 'v1.0.1' / '1.0' 规整成 (主,次,修订) 三元组，便于比较"""
    nums = re.findall(r"\d+", v or "")
    nums = (nums + ["0", "0", "0"])[:3]
    return tuple(int(x) for x in nums)


def _ssl_ctx():
    """用 certifi 的 CA 证书包构建 SSL 上下文，避免冻结 exe 在缺根证书的机器上
    报 CERTIFICATE_VERIFY_FAILED（unable to get local issuer certificate）。"""
    return ssl.create_default_context(cafile=certifi.where())


def check_update() -> dict:
    """检测 Gitee 仓库最新 Release。

    返回:
      {"has_update": bool, "latest": str, "url": str, "notes": str,
       "download_url": str}  # exe 直接下载链接（若有）
      仓库未配置 / 出错时返回 {"has_update": False, "skipped": True} 或 {"has_update": False, "error": str}
    """
    if GITEE_OWNER in ("", "你的Gitee用户名"):
        return {"has_update": False, "skipped": True}
    try:
        url = GITEE_API
        if GITEE_TOKEN:
            url += "?access_token=" + GITEE_TOKEN
        req = urllib.request.Request(
            url, headers={"User-Agent": "CNKI-Citation-Tool"})
        with urllib.request.urlopen(req, timeout=8, context=_ssl_ctx()) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        latest = (data.get("tag_name") or data.get("name") or "").strip()
        if not latest:
            return {"has_update": False}
        has = _norm_ver(latest) > _norm_ver(APP_VERSION)

        # 尝试获取 exe 附件的直接下载链接
        download_url = ""
        release_id = data.get("id")
        if release_id and GITEE_TOKEN:
            try:
                att_url = (f"https://gitee.com/api/v5/repos/{GITEE_OWNER}/{GITEE_REPO}"
                           f"/releases/{release_id}/attach_files"
                           f"?access_token={GITEE_TOKEN}")
                att_req = urllib.request.Request(att_url,
                                                headers={"User-Agent": "CNKI-Citation-Tool"})
                with urllib.request.urlopen(att_req, timeout=10, context=_ssl_ctx()) as att_resp:
                    attachments = json.loads(att_resp.read().decode("utf-8"))
                if isinstance(attachments, list):
                    for a in attachments:
                        name = (a.get("name") or "")
                        # 精确匹配当前版本的 exe 名（避免下载到另一个平台的 exe）
                        if name == f"{APP_NAME}.exe" or name.endswith(f"{APP_NAME}.exe"):
                            aid = a.get("id")
                            # 私有仓库的 browser_download_url 不带 token 会 403，
                            # 改用 Gitee API 附件下载接口（带 access_token 可正常鉴权）。
                            download_url = (
                                f"https://gitee.com/api/v5/repos/{GITEE_OWNER}/{GITEE_REPO}"
                                f"/releases/{release_id}/attach_files/{aid}/download"
                                f"?access_token={GITEE_TOKEN}")
                            break
                    # 兜底：精确匹配没命中时，仍尝试取第一个 .exe（兼容旧版 Release 只有一个 exe 的情况）
                    if not download_url:
                        for a in attachments:
                            name = (a.get("name") or "").lower()
                            if name.endswith(".exe"):
                                aid = a.get("id")
                                download_url = (
                                    f"https://gitee.com/api/v5/repos/{GITEE_OWNER}/{GITEE_REPO}"
                                    f"/releases/{release_id}/attach_files/{aid}/download"
                                    f"?access_token={GITEE_TOKEN}")
                                break
            except Exception:
                pass  # 下载链接拿不到不影响基本检测

        return {
            "has_update": has,
            "latest": latest,
            "url": data.get("html_url") or "",
            "notes": (data.get("body") or "").strip(),
            "download_url": download_url,
        }
    except Exception as e:
        return {"has_update": False, "error": str(e)}



# 全局日志回调（GUI 可设置；为 None 时走标准输出）
_on_log = None


def set_log_callback(cb):
    """设置日志回调：cb(text: str) -> None。GUI 模式用于实时更新文本框"""
    global _on_log
    _on_log = cb


def log(tag: str, msg: str = ""):
    text = f"[{tag}] {msg}" if msg else f"[{tag}]"
    if _on_log:
        _on_log(text)
    else:
        print(text)


async def user_confirm(prompt: str = "完成后按回车继续..."):
    """在事件循环外等待用户输入"""
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, input, prompt)


class HumanBehavior:
    """拟人行为模拟"""

    @staticmethod
    async def pause(min_s: float = 0.4, max_s: float = 1.2):
        await asyncio.sleep(random.uniform(min_s, max_s))

    @staticmethod
    def _bezier(t: float, p0, p1, p2, p3):
        mt = 1 - t
        return (
            mt**3 * p0[0] + 3 * mt**2 * t * p1[0] + 3 * mt * t**2 * p2[0] + t**3 * p3[0],
            mt**3 * p0[1] + 3 * mt**2 * t * p1[1] + 3 * mt * t**2 * p2[1] + t**3 * p3[1],
        )

    async def move_mouse_to(self, page: Page, x: float, y: float, steps: int = None):
        """三次贝塞尔曲线鼠标移动 + 轻微抖动"""
        cur = await page.evaluate("() => ({x: window.__mx || 0, y: window.__my || 0})")
        cur = cur or {"x": 0, "y": 0}
        sx, sy = float(cur["x"]), float(cur["y"])
        steps = steps or random.randint(15, 32)

        cp1 = (sx + random.uniform(-120, 120), sy + random.uniform(-90, 90))
        cp2 = (x + random.uniform(-90, 90), y + random.uniform(-70, 70))

        for i in range(1, steps + 1):
            t = i / steps
            px, py = self._bezier(t, (sx, sy), cp1, cp2, (x, y))
            if random.random() < 0.22:
                px += random.uniform(-1.5, 1.5)
                py += random.uniform(-1.5, 1.5)
            await page.mouse.move(px, py)
            await asyncio.sleep(random.uniform(0.004, 0.018))

        await page.evaluate("(p) => { window.__mx = p.x; window.__my = p.y; }", {"x": x, "y": y})

    async def click_element(self, page: Page, element):
        """先移动到元素（拟人轨迹）再点击"""
        try:
            box = await element.bounding_box()
        except Exception:
            box = None

        if not box:
            await element.click()
            return

        tx = box["x"] + box["width"] * random.uniform(0.35, 0.65)
        ty = box["y"] + box["height"] * random.uniform(0.35, 0.65)
        await self.move_mouse_to(page, tx, ty)
        await self.pause(0.08, 0.3)
        await page.mouse.click(tx, ty)

    async def human_type(self, page: Page, element, text: str):
        """逐字输入，随机延迟，偶发思考停顿"""
        await self.click_element(page, element)
        await self.pause(0.2, 0.5)

        # 清空已有内容
        await page.keyboard.press("Control+A")
        await page.keyboard.press("Delete")
        await self.pause(0.1, 0.25)

        for ch in text:
            await page.keyboard.type(ch)
            delay = random.uniform(0.05, 0.15)
            if random.random() < 0.07:          # 7% 概率长停顿（模拟思考）
                delay += random.uniform(0.2, 0.5)
            await asyncio.sleep(delay)

    async def random_scroll(self, page: Page, rounds: int = None):
        """随机滚动，模拟浏览阅读"""
        for _ in range(rounds or random.randint(2, 4)):
            await page.mouse.wheel(0, random.randint(120, 420))
            await self.pause(0.25, 0.7)


async def start_browser(p, connect: bool = False, cdp: str = "http://localhost:9222",
                        human: bool = True) -> tuple:
    """启动浏览器。返回 (context, page, need_close)"""
    if connect:
        log("1/4", f"连接已有 Chrome ({cdp})")
        browser = await p.chromium.connect_over_cdp(cdp)
        context = browser.contexts[0] if browser.contexts else await browser.new_context()
        page = context.pages[0] if context.pages else await context.new_page()
        await Stealth().apply_stealth_async(page)
        return context, page, False

    log("1/4", "启动浏览器（系统 Chrome + 持久化 Profile）")

    common = dict(
        headless=False,
        viewport={"width": 1920, "height": 1080},   # 固定视口确保布局正常
        locale="zh-CN",
        timezone_id="Asia/Shanghai",
        ignore_default_args=["--enable-automation"],   # 移除自动化标志
        args=[
            "--disable-blink-features=AutomationControlled",
            "--disable-infobars",
            "--no-sandbox",
            "--window-size=1920,1080",
        ],
    )

    context = None
    # 优先使用系统真实 Chrome（指纹最真实）
    try:
        context = await p.chromium.launch_persistent_context(
            user_data_dir=str(PROFILE_DIR), channel="chrome", **common
        )
        log("   ", "已使用系统 Google Chrome")
    except Exception as e:
        log("   ", f"系统 Chrome 不可用（{e}），回退到内置 Chromium")
        context = await p.chromium.launch_persistent_context(
            user_data_dir=str(PROFILE_DIR), **common
        )

    page = context.pages[0] if context.pages else await context.new_page()
    await Stealth().apply_stealth_async(page)
    return context, page, True


async def check_block(page: Page) -> bool:
    """检测是否被拦截 / 需要验证码 / 需要登录。返回 True 表示需要人工介入"""
    await asyncio.sleep(1.2)
    url = (page.url or "").lower()
    try:
        title = await page.title() or ""
    except Exception:
        title = ""

    need = (
        any(k in url for k in ("verify", "captcha", "valid", "login", "passport"))
        or "验证" in title or "403" in title or "拒绝" in title or "登录" in title
    )

    if need:
        log("⚠ 需人工介入", "请在浏览器窗口中完成登录 / 验证码")
        await user_confirm("完成后按回车继续...")
        await asyncio.sleep(1.5)
        return True
    return False


async def search_paper(page: Page, title: str, H: Optional[HumanBehavior]) -> bool:
    """搜索论文。优先用直接 URL（最可靠），H 为 None 时跳过拟人"""
    log("2/4", f"搜索：{title}")

    await check_block(page)

    # ── 策略 A：直接构造搜索 URL（绕过首页，100% 可靠）──
    kw = urllib.parse.quote(title)
    # kns8s 是 CNKI 新版搜索系统；korder=TI 表示按标题匹配
    search_url = f"https://kns.cnki.net/kns8s/defaultresult/index?kw={kw}&korder=TI"

    log("   ", f"直接导航到搜索页...")
    await page.goto(search_url, wait_until="domcontentloaded", timeout=60000)

    if H:
        await asyncio.sleep(random.uniform(1, 2))
        await H.random_scroll(page, rounds=random.randint(1, 2))
    else:
        await asyncio.sleep(3)

    try:
        await page.wait_for_load_state("networkidle", timeout=30000)
    except Exception:
        pass
    await asyncio.sleep(2)

    log("   ", f"当前 URL: {page.url}")

    await check_block(page)

    return True


async def click_cite(page: Page, title: str, H: Optional[HumanBehavior]) -> Page:
    """在结果中定位论文并点击引用按钮，返回引用页"""
    log("3/4", "定位论文并点击「引用」")

    try:
        await page.wait_for_selector(".result-list, table.result-grid-list, #gridTable", timeout=12000)
    except Exception:
        pass

    if H:
        await H.random_scroll(page, rounds=random.randint(1, 2))

    short = title[:15]

    link = None
    # 排除筛选/侧边栏区域，只匹配真正搜索结果中的标题链接
    exclude_selectors = [
        ".conditions", ".facet", ".filter", ".sidebar", ".left-panel",
        "#facet-group", ".cluster", "[class*='filter']", "[class*='facet']",
        "[class*='cluster']"
    ]
    for sel in (
        # 右侧结果区的标题链接（按优先级）
        "#gridTable .name a",                  # 经典表格视图
        "table.result-grid-list .fz14",         # 网格列表
        ".result-table-list .name a",           # 新版列表
        "#resultTable .name a",
        ".right-layout .result-list a.fz14",   # 右侧布局
        f"a.fz14:has-text('{short}')",         # 通用：带 fz14 类的标题
        f".result-content a:has-text('{short}')",
    ):
        try:
            el = await page.query_selector(sel)
            if el and await el.is_visible():
                # 二次确认不在排除区域内
                parent_html = (await el.evaluate("el => el.parentElement?.outerHTML || ''") or "")
                if not any(ex in parent_html for ex in ["class=\"conditions\"", "class=\"facet\""]):
                    link = el
                    log("   ", f"找到论文链接: {sel}")
                    break
        except Exception:
            continue

    if not link:
        # 调试：dump 页面上所有链接
        all_links = await page.evaluate("""() => {
            const links = Array.from(document.querySelectorAll('a'));
            return links.slice(0, 30).map(a => ({
                text: (a.textContent || '').trim().slice(0, 50),
                href: (a.href || '').slice(0, 100),
                cls: a.className || '',
                title: a.title || ''
            }));
        }""")
        log("   ", f"页面上前 {len(all_links)} 个链接:")
        for i, a in enumerate(all_links[:10]):
            log("   ", f"  [{i}] {a.get('text','')} | {a.get('href','')[:60]} | {a.get('cls','')[:30]}")

        shot = OUTPUT_DIR / "search_results.png"
        await page.screenshot(path=str(shot))
        raise RuntimeError(f"未找到论文「{title}」，截图已保存：{shot}")

    # 在标题所在行/区域内查找引用按钮
    # 先调试：dump 论文链接信息和附近所有链接
    debug_info = await page.evaluate(
        """(el) => {
            const result = {
                linkHref: el.href || '',
                linkText: (el.textContent || '').trim().slice(0, 80),
                linkClass: el.className || '',
                parentId: el.parentElement ? el.parentElement.tagName : '',
                parentClass: el.parentElement ? (el.parentElement.className || '').slice(0, 60) : '',
                nearbyLinks: []
            };

            // 向上找到行容器（最多5层）
            let row = el.parentElement;
            for (let i = 0; i < 6 && row; i++) {
                const tag = row.tagName;
                const cls = (row.className || '').toString();
                result.nearbyLinks.push({
                    level: i,
                    tag: tag,
                    cls: cls.slice(0, 60),
                    linkCount: row.querySelectorAll('a').length,
                    allHrefs: Array.from(row.querySelectorAll('a')).map(a => ({
                        href: (a.href||'').slice(0,120),
                        text: (a.textContent||'').trim().slice(0,30),
                        title: a.title||'',
                        cls: (a.className||'').slice(0,30)
                    }))
                });
                if (tag === 'TR' || cls.includes('single') || cls.includes('item') ||
                    cls.includes('result') || cls.includes('table')) break;
                row = row.parentElement;
            }
            return result;
        }""",
        link,
    )

    log("   ", f"论文链接: {debug_info.get('linkHref','')}")
    log("   ", f"父元素: <{debug_info.get('parentId','')} class=\"{debug_info.get('parentClass','')}\">")

    # 查找引用按钮——返回元素信息（CNKI 的引用按钮可能是 javascript:void(0) + onclick）
    cite_btn_info = await page.evaluate(
        """(el) => {
            let row = el;
            for (let i = 0; i < 6 && row; i++) {
                const tag = row.tagName;
                const cls = (row.className || '').toString();
                if (tag === 'TR' || cls.includes('single') || cls.includes('item') ||
                    cls.includes('result') || cls.includes('table')) break;
                row = row.parentElement;
            }
            const scope = row || el.parentElement;
            const links = scope ? scope.querySelectorAll('a') : [];
            for (const a of links) {
                const href = a.href || '';
                const t = (a.title || '') + (a.textContent || '');
                if (href.includes('cite') || href.includes('refer') ||
                    t.includes('引用') || href.includes('citation')) {
                    return {
                        href: href,
                        isJsLink: href.startsWith('javascript:') || !href,
                        text: (a.textContent || '').trim().slice(0, 20),
                        title: a.title || '',
                        className: a.className || '',
                        // 用于重新定位元素的线索
                        parentTag: a.parentElement ? a.parentElement.tagName : '',
                        parentClass: a.parentElement ? (a.parentElement.className || '').slice(0, 40) : '',
                    };
                }
            }
            return null;
        }""",
        link,
    )

    if not cite_btn_info:
        shot = OUTPUT_DIR / "no_cite_button.png"
        await page.screenshot(path=str(shot))
        raise RuntimeError(f"找到论文但未定位到「引用」按钮，截图已保存：{shot}")

    log("   ", f"引用按钮: href={cite_btn_info.get('href','')[:60]} js={cite_btn_info.get('isJsLink')} "
              f"text={cite_btn_info.get('text','')}")

    context = page.context

    # 根据按钮类型选择操作方式
    if cite_btn_info.get("isJsLink"):
        # JS 按钮（javascript:void(0)）→ 直接点击 DOM 元素
        log("   ", "使用 click() 点击 JS 引用按钮...")
        cite_page = None

        # 先尝试通过多种方式点击引用按钮
        btn_text = cite_btn_info.get("text", "").strip()
        btn_class = cite_btn_info.get("className", "").split()[0] if cite_btn_info.get("className") else ""
        clicked = False

        # 策略1：JS 在目标论文行内直接查找并点击引用链接
        if not clicked:
            try:
                result = await page.evaluate("""(titleText) => {
                    const rows = document.querySelectorAll('#gridTable tr, table.result-grid-list tr');
                    for (const row of rows) {
                        // 确认这行包含目标论文
                        const nameLink = row.querySelector('.name a, td.name a, a.fz14');
                        if (!nameLink || !nameLink.textContent.includes(titleText)) continue;

                        // 在该行中找引用按钮
                        for (const a of row.querySelectorAll('a')) {
                            const h = (a.href || '').toLowerCase();
                            const t = (a.title || '') + (a.textContent || '');
                            const img = a.querySelector('img');
                            const src = img ? (img.src || '').toLowerCase() : '';
                            if (h.includes('cite') || t.includes('引用') || src.includes('cite') || src.includes('quote')) {
                                a.click();
                                return {clicked: true, tag: a.tagName, href: a.href, text: a.textContent.trim().slice(0,20)};
                            }
                        }
                    }
                    return {clicked: false};
                }""", title[:15])
                clicked = result.get("clicked", False) if result else False
                if clicked:
                    log("   ", f"JS 点击成功: {result}")
            except Exception as e:
                log("   ", f"JS 点击异常: {e}")

        # 等待新页面或弹窗出现
        await asyncio.sleep(3)

        # 检查是否有新标签页
        pages = context.pages
        if len(pages) > 1:
            cite_page = pages[-1]
            log("   ", "检测到新标签页")
        else:
            # 没有新标签页，可能是弹窗——使用当前页面并截图
            cite_page = page
            log("   ", "无新标签页，检查当前页面是否出现弹窗...")
            await page.screenshot(path=str(OUTPUT_DIR / "debug_after_cite_click.png"))

        try:
            await cite_page.wait_for_load_state("networkidle", timeout=10000)
        except Exception:
            pass
        await asyncio.sleep(1.5)
    else:
        # 普通 URL 链接 → window.open
        log("   ", "使用 window.open 打开引用页...")
        async with context.expect_page(timeout=15000) as new_page_info:
            await page.evaluate("(h) => window.open(h, '_blank')", cite_btn_info["href"])
        cite_page = await new_page_info.value

    try:
        await cite_page.wait_for_load_state("networkidle", timeout=20000)
    except Exception:
        pass
    await asyncio.sleep(1.5)
    return cite_page


async def extract_citation(cite_page: Page) -> Optional[str]:
    """提取 GB/T 7714-2025 格式引文"""
    log("4/4", "提取 GB/T 7714-2025 格式引文")

    text = None

    # 1) textarea
    for ta in await cite_page.query_selector_all("textarea"):
        v = (await ta.input_value() or "").strip()
        if "7714" in v or ("[J]" in v and len(v) > 40):
            text = v
            break

    # 2) 页面文本正则
    if not text:
        try:
            body = (await cite_page.evaluate("() => document.body?.innerText || ''") or "")
        except Exception:
            body = ""
        for pat in (
            # CNKI 特有：多种格式连在一起，GB/T 7714 在 "格式引文" 和 "MLA" 之间
            r'格式引文\s*\[?\d+\]?\s*(.+?)(?=\s*MLA格式引文|\s*APA格式引文|\s*知网研学)',
            r'GB/T\s*7714[—\-–]?\s*2025[：:\s]*\n?(.*?)(?=\n\s*\n|\n[A-Za-z]{2,}\s*7714|\Z)',
            r'\[1\]\s*(.*?)(?=\n\s*\[2\]|\Z)',
        ):
            m = re.search(pat, body, re.DOTALL)
            if m and len(m.group(1).strip()) > 25:
                text = m.group(1).strip()
                break

    # 3) 指定容器
    if not text:
        for sel in (".citation-content", ".ref-content", "[class*='citation']",
                    "[class*='reference']", "#copyText"):
            el = await cite_page.query_selector(sel)
            if el:
                try:
                    v = (await el.evaluate("el => el.innerText") or "").strip()
                except Exception:
                    v = ""
                if len(v) > 25:
                    text = v
                    break

    if not text:
        return None

    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"^\[\d+\]\s*", "", text)
    return text


async def process_one(page: Page, title: str, human: bool) -> dict:
    """处理单篇论文"""
    H = HumanBehavior() if human else None
    rec = {"title": title, "citation": None, "success": False, "error": None}

    try:
        await search_paper(page, title, H)
        cite_page = await click_cite(page, title, H)
        is_new_page = (cite_page is not page)   # 仅新标签页需关闭，避免误关主页
        try:
            citation = await extract_citation(cite_page)
        finally:
            if is_new_page:
                await cite_page.close()

        if citation and len(citation) > 20:
            rec["citation"] = citation
            rec["success"] = True
            log("✓ 成功", citation)
        else:
            raise RuntimeError("未能提取到有效引文内容")

    except Exception as e:
        rec["error"] = str(e)
        log("✗ 失败", str(e))
        try:
            shot = OUTPUT_DIR / "error.png"
            await page.screenshot(path=str(shot))
        except Exception:
            pass

    return rec


def _resolve_col(ws, spec: str, header_row: int = 1) -> int:
    """解析列引用：字母(A/B/C) / 数字(1-based) / 列名(按表头查找)"""
    spec = str(spec).strip()
    # 纯 ASCII 字母（A-Z）→ 列字母；中文等非 ASCII 走列名查找
    if spec.isascii() and spec.isalpha():
        return column_index_from_string(spec.upper())
    try:
        return int(spec)
    except ValueError:
        pass
    for cell in ws[header_row]:
        if cell.value is not None and str(cell.value).strip() == spec:
            return cell.column
    raise ValueError(f"未找到列：{spec}")


def load_excel(path: str, title_col: str, out_col: str):
    """加载 Excel，返回 (wb, ws, title_idx, out_idx, tasks)
    tasks: [(row_idx, title), ...]
    若 out_col 不存在则自动新建（追加列 + 写表头）
    """
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        raise ValueError(
            f"无法读取 Excel 文件「{Path(path).name}」：本工具用 openpyxl 直接解析 .xlsx，"
            f"无需安装 Office。请确认：① 文件是 .xlsx 格式（旧版 .xls 不被支持，"
            f"请用 Excel/WPS 另存为 .xlsx）；② 文件未损坏。\n底层错误：{e}"
        )
    ws = wb.active
    t_idx = _resolve_col(ws, title_col)
    try:
        o_idx = _resolve_col(ws, out_col)
    except ValueError:
        o_idx = ws.max_column + 1
        ws.cell(row=1, column=o_idx, value=out_col)
    tasks = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=t_idx).value
        if v and str(v).strip():
            tasks.append((r, str(v).strip()))
    return wb, ws, t_idx, o_idx, tasks


async def run(titles: List[str], connect: bool, cdp: str, human: bool,
              out_prefix: str = "citations", excel_path: Optional[str] = None,
              title_col: str = "标题", out_col: str = "引文",
              min_gap: float = 5.0, max_gap: float = 12.0,
              on_log=None) -> List[dict]:
    # GUI 模式：设置日志回调（实时更新界面）
    if on_log:
        set_log_callback(on_log)

    # Excel 模式：加载并覆盖 titles / tasks
    wb = ws = t_idx = o_idx = None
    tasks = None
    if excel_path:
        wb, ws, t_idx, o_idx, tasks = load_excel(excel_path, title_col, out_col)
        titles = [t for _, t in tasks]
        log("   ", f"Excel 加载完成：共 {len(titles)} 条待处理")

    results = []
    async with async_playwright() as p:
        context, page, need_close = await start_browser(p, connect, cdp, human)
        try:
            await page.goto(CNKI_HOME, wait_until="domcontentloaded", timeout=60000)
            await asyncio.sleep(2)
            await check_block(page)

            for i, t in enumerate(titles, 1):
                log("进度", f"第 {i}/{len(titles)} 篇：{t}")
                row_idx = tasks[i - 1][0] if tasks else None

                # 断点续传：跳过已回填的行
                if excel_path and row_idx:
                    existing = ws.cell(row=row_idx, column=o_idx).value
                    if existing and str(existing).strip():
                        log("   ", "该行列已存在引文，跳过（断点续传）")
                        results.append({"title": t, "citation": str(existing).strip(),
                                        "success": True, "error": None, "skipped": True})
                        continue

                rec = await process_one(page, t, human)
                results.append(rec)

                # 回填 Excel（成功后实时保存，防丢失）
                if excel_path and row_idx and rec.get("success") and rec.get("citation"):
                    ws.cell(row=row_idx, column=o_idx, value=rec["citation"])
                    wb.save(excel_path)
                    log("   ", f"已回填 Excel 第 {row_idx} 行并保存")

                # 间隔控制（规避 CNKI 兜底速率校验）
                if i < len(titles):
                    gap = random.uniform(min_gap, max_gap)
                    if random.random() < 0.15:   # 偶发长间隔，更像真人翻页/思考
                        gap += random.uniform(3, 8)
                    log("   ", f"间隔 {gap:.1f}s 后继续...")
                    await asyncio.sleep(gap)
        finally:
            if need_close:
                await context.close()

    # 保存 txt / json 备份
    ok = [r for r in results if r["success"]]
    lines = [f"# CNKI 引文导出（GB/T 7714-2025）\n# 导出时间：{datetime.now():%Y-%m-%d %H:%M:%S}\n"]
    for i, r in enumerate(ok, 1):
        lines.append(f"[{i}] {r['citation']}\n")
    if len(ok) < len(results):
        lines.append("\n# 以下篇目提取失败：\n")
        for r in results:
            if not r["success"]:
                lines.append(f"- {r['title']}  ->  {r['error']}\n")

    txt = OUTPUT_DIR / f"{out_prefix}.txt"
    txt.write_text("\n".join(lines), encoding="utf-8")

    js = OUTPUT_DIR / f"{out_prefix}.json"
    js.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")

    skipped = sum(1 for r in results if r.get("skipped"))
    real_ok = len(ok) - skipped
    log("完成", f"本次成功 {real_ok} 条，跳过(已存在) {skipped} 条，失败 {len(results) - len(ok)} 条")
    log("输出", f"文本：{txt}")
    log("输出", f"JSON：{js}")
    if excel_path:
        log("输出", f"Excel：{excel_path}（已实时回填）")
    return results


def main():
    ap = argparse.ArgumentParser(description="CNKI 论文引文获取工具（GB/T 7714-2025）")
    ap.add_argument("title", nargs="*", help="论文标题")
    ap.add_argument("-f", "--file", help="批量文件：每行一个标题")
    ap.add_argument("--excel", help="Excel 模式：解析该 xlsx，按标题列逐条搜索，引文回填到引文列")
    ap.add_argument("--title-col", default="标题", help="Excel 标题列（列名/字母/数字），默认 标题")
    ap.add_argument("--out-col", default="引文", help="Excel 引文列（不存在则自动新建），默认 引文")
    ap.add_argument("--connect", action="store_true", help="连接已打开的 Chrome（--remote-debugging-port=9222）")
    ap.add_argument("--cdp", default="http://localhost:9222", help="CDP 地址")
    ap.add_argument("--fast", action="store_true", help="关闭拟人行为（快速但更易触发校验）")
    ap.add_argument("-o", "--out", default="citations", help="输出文件名前缀")
    ap.add_argument("--min-gap", type=float, default=5.0, help="搜索间隔下限(秒)，默认 5")
    ap.add_argument("--max-gap", type=float, default=12.0, help="搜索间隔上限(秒)，默认 12")
    args = ap.parse_args()

    titles = []
    if args.excel:
        fp = Path(args.excel)
        if not fp.exists():
            print(f"Excel 文件不存在：{fp}")
            return 1
        print(f"Excel 模式：{fp}")
        print(f"  标题列={args.title_col}  引文列={args.out_col}")
        print(f"  间隔={args.min_gap}-{args.max_gap}s（规避速率校验）")
        # titles 在 run() 内从 Excel 加载
    elif args.file:
        fp = Path(args.file)
        if not fp.exists():
            print(f"文件不存在：{fp}")
            return 1
        titles = [l.strip() for l in fp.read_text(encoding="utf-8").splitlines()
                  if l.strip() and not l.strip().startswith("#")]
    elif args.title:
        titles = [" ".join(args.title)]
    else:
        titles = ["新形势下企业财务管理信息化建设的途径研究"]
        print(f"未指定标题，使用示例：{titles[0]}")
        print("用法：python cnki_citation.py \"标题\"  |  python cnki_citation.py -f titles.txt\n")

    return asyncio.run(run(
        titles, args.connect, args.cdp, not args.fast, args.out,
        excel_path=args.excel, title_col=args.title_col, out_col=args.out_col,
        min_gap=args.min_gap, max_gap=args.max_gap,
    ))


if __name__ == "__main__":
    raise SystemExit(0 if not main() else 0)
